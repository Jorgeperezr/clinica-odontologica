"""
Exportación de reportes a Excel (RF-REP-06).
Genera el .xlsx en memoria con openpyxl y lo devuelve como descarga.
Son datos ya calculados (no fórmulas), así que no requiere recálculo.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def build_xlsx(title, headers, rows):
    """
    Construye un .xlsx en memoria con un encabezado con estilo y las filas.
    Devuelve los bytes listos para una HttpResponse de descarga.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limita el nombre de hoja a 31 chars

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2563EB")

    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        ws.append(row)

    # Ancho de columnas automático (aproximado)
    for col, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(r[col - 1])) for r in rows]) if rows else len(str(header))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
