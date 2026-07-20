"""
Autocompletado del formulario oficial MSP HCU-form.033/2021 (Sprint 31).

Carga la plantilla oficial del Ministerio (xlsx, tal cual se publica) y
escribe en sus celdas los datos registrados en el sistema: datos del
paciente, motivo, enfermedad actual, constantes vitales, índices CPO-ceo,
diagnósticos CIE-10 y datos del profesional. El resultado es el documento
oficial listo para imprimir o archivar, sin retipear nada.

Las coordenadas provienen del mapeo de la plantilla oficial. Si el MSP
publica una versión nueva del formulario, hay que revisar este mapeo.
"""

import io
from pathlib import Path

import openpyxl

TEMPLATE_PATH = Path(__file__).resolve().parent / "resources" / "hcu_form033.xlsx"


def _set(ws, coord, value):
    """
    Escribe respetando celdas combinadas (en la esquina superior izquierda).
    En esta plantilla varias celdas combinan ETIQUETA y campo en un mismo
    rango (ej. "PRIMER APELLIDO" en A4:L5): en ese caso el valor se anexa
    en una línea nueva bajo la etiqueta, sin destruirla.
    """
    if value in (None, ""):
        return
    cell = ws[coord]
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            cell = ws.cell(row=merged.min_row, column=merged.min_col)
            break
    existing = cell.value
    if existing not in (None, ""):
        cell.value = f"{existing}\n{value}"
        try:
            cell.alignment = cell.alignment.copy(wrap_text=True)
        except Exception:
            pass
    else:
        cell.value = value


def build_form033_xlsx(patient, tenant, form_record, diagnoses, cpo_ceo, professional):
    """Devuelve los bytes del formulario oficial con los datos del sistema."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws1, ws2 = wb["1"], wb["2"]

    # ── Hoja 1 ──
    # A. Establecimiento y paciente
    _set(ws1, "T3", tenant.name)
    apellidos = (patient.last_name or "").split(" ", 1)
    nombres = (patient.first_name or "").split(" ", 1)
    _set(ws1, "A5", apellidos[0] if apellidos else "")
    _set(ws1, "M5", apellidos[1] if len(apellidos) > 1 else "")
    _set(ws1, "Y5", nombres[0] if nombres else "")
    _set(ws1, "AK5", nombres[1] if len(nombres) > 1 else "")
    _set(ws1, "AW5", getattr(patient, "sex", "") or "")

    if form_record:
        # B y C
        motivo = form_record.motivo_consulta or ""
        if form_record.embarazada is not None:
            emb = "SI" if form_record.embarazada else "NO"
            motivo = f"{motivo}     |     EMBARAZADA: {emb}" if motivo else f"EMBARAZADA: {emb}"
        _set(ws1, "A9", motivo)
        _set(ws1, "A12", form_record.enfermedad_actual)
        # F. Constantes vitales (fila de datos bajo las etiquetas de la 33)
        _set(ws1, "A34", form_record.temperatura)
        _set(ws1, "Q34", form_record.pulso)
        _set(ws1, "AH34", form_record.frecuencia_respiratoria)
        _set(ws1, "AV34", form_record.presion_arterial)

    # J. Índices CPO-ceo (los TOTALES son fórmulas de la plantilla: se respetan)
    cpo, ceo = cpo_ceo["cpo"], cpo_ceo["ceo"]
    _set(ws1, "AZ68", cpo["C"])
    _set(ws1, "BC68", cpo["P"])
    _set(ws1, "BF68", cpo["O"])
    _set(ws1, "AZ72", ceo["c"])
    _set(ws1, "BC72", ceo["e"])
    _set(ws1, "BF72", ceo["o"])

    # ── Hoja 2 ──
    # N. Diagnósticos: filas 16-18 (1-3 izquierda, 4-6 derecha)
    left_rows, right_rows = [16, 17, 18], [16, 17, 18]
    for i, d in enumerate(diagnoses[:6]):
        if i < 3:
            r = left_rows[i]
            _set(ws2, f"B{r}", d.description)
            _set(ws2, f"Y{r}", d.code)
            _set(ws2, f"AC{r}" if d.diagnosis_kind == "pre" else f"AE{r}", "X")
        else:
            r = right_rows[i - 3]
            _set(ws2, f"AH{r}", d.description)
            _set(ws2, f"BE{r}", d.code)
            _set(ws2, f"BI{r}" if d.diagnosis_kind == "pre" else f"BK{r}", "X")

    # O. Profesional (automático desde la credencial)
    recorded = professional.get("recorded_at", "")
    full_name = (professional.get("full_name") or "").split(" ")
    _set(ws2, "A22", recorded[:10])
    _set(ws2, "I22", recorded[11:16])
    _set(ws2, "O22", full_name[0] if full_name else "")
    _set(ws2, "AH22", full_name[1] if len(full_name) > 1 else "")
    _set(ws2, "AY22", full_name[2] if len(full_name) > 2 else "")
    _set(ws2, "A24", professional.get("license_number") or "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
